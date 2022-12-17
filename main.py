import pandas as pd
import openpyxl
import os
import sys


def validate_argv(argv):
    if argv is not None:
        if len(argv) == 3:
            return True
        else:
            print("Please enter data folder path and if you want to reprocess the files")
            return False
    else:
        print("No Arguments")
        return False


def remove_nan_from_df(df):
    df = df.dropna(how='all')
    return df


def read_excels_to_df(path):
    processed_path = path + '\\processed\\'
    files = os.listdir(processed_path)
    files_dict = {}
    for file in files:
        if file.endswith(".xlsx") and file != "START.xlsx":
            file_dfs = pd.read_excel(processed_path + file, None)
            files_dict[file.split('.')[0]] = file_dfs

    return files_dict


def read_start_file(files_dfs):
    if files_dfs['START.xlsx'].empty:
        print("START.xlsx is empty")
        return None
    else:
        print(files_dfs['START.xlsx'].head())


def does_sheet_row_only_have_value_in_first_column(row):
    # iterate over all cells in row
    for cell in row:
        if cell.value is not None and cell.column != 1:
            return False
    return True


def find_all_sheet_rows_with_bolded_value_in_first_column(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=10):
        if row[0].font.bold:
            rows.append(row[0].row)

    return rows


def find_all_sheet_rows_with_value_in_only_first_column(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=10):
        if does_sheet_row_only_have_value_in_first_column(row):
            rows.append(row[0].row)

    return rows


def copy_column_to_other(sheet, src_col, dest_col):
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=src_col, max_col=src_col):
        if sheet.cell(row=row[0].row, column=dest_col).value is None:
            sheet.cell(row=row[0].row, column=dest_col).value = row[0].value


def remove_empty_rows(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value is None:
            sheet.delete_rows(row[0].row, 1)


def unmerge_whole_sheet(sheet):
    for merge in list(sheet.merged_cells):
        sheet.unmerge_cells(range_string=str(merge))


def prep_balance_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet):
    unmerge_whole_sheet(sheet)
    sheet.delete_rows(1, 1)
    sheet.cell(row=1, column=1).value = "PARAMETERS"
    remove_empty_rows(sheet)

    # copy all values from column C to column B
    copy_column_to_other(sheet, 3, 2)
    copy_column_to_other(sheet, 5, 4)
    copy_column_to_other(sheet, 7, 6)
    sheet.delete_cols(3, 1)
    sheet.delete_cols(4, 1)
    sheet.delete_cols(5, 1)
    sheet.delete_rows(27, 1)

    first_column_only_rows = find_all_sheet_rows_with_value_in_only_first_column(sheet)
    first_column_bolded_rows = find_all_sheet_rows_with_bolded_value_in_first_column(sheet)
    copy_first_only_values_to_following_rows(sheet, first_column_only_rows, first_column_bolded_rows)
    delete_rows(first_column_only_rows, sheet)


def delete_rows(rows, sheet):
    for i in range(len(rows)):
        sheet.delete_rows(rows[i] - i, 1)


def copy_first_only_values_to_following_rows(sheet, first_column_only_rows, first_column_bolded_rows):
    for row in first_column_only_rows:
        bolded_index = first_column_bolded_rows.index(row)
        if bolded_index + 1 < len(first_column_bolded_rows):
            for i in range(row + 1, first_column_bolded_rows[bolded_index + 1]):
                sheet.cell(row=i, column=1).value = sheet.cell(row=row, column=1).value + " " + sheet.cell(row=i,
                                                                                                           column=1).value


def create_country_header(sheet, start_column, has_third_column=False):
    country = sheet.cell(row=1, column=start_column).value
    sheet.cell(row=1, column=start_column).value = country + " " + sheet.cell(row=2, column=start_column).value
    next_value = country + " " + sheet.cell(row=2, column=start_column + 1).value
    if has_third_column:
        second_value = next_value + " 1"
        third_value = next_value + " 2"
        sheet.cell(row=1, column=start_column + 1).value = second_value
        sheet.cell(row=1, column=start_column + 2).value = third_value
    else:
        sheet.cell(row=1, column=start_column + 1).value = next_value


def prep_income_statement(sheet: openpyxl.worksheet.worksheet.Worksheet):
    unmerge_whole_sheet(sheet)
    sheet.delete_rows(1, 1)
    sheet.cell(row=1, column=1).value = "PARAMETERS"
    remove_empty_rows(sheet)
    create_country_header(sheet, 2)
    create_country_header(sheet, 4)
    create_country_header(sheet, 6)
    sheet.delete_rows(2, 1)
    first_column_only_rows = find_all_sheet_rows_with_value_in_only_first_column(sheet)
    first_column_bolded_rows = find_all_sheet_rows_with_bolded_value_in_first_column(sheet)
    copy_first_only_values_to_following_rows(sheet, first_column_only_rows, first_column_bolded_rows)
    delete_rows(first_column_only_rows, sheet)


def prep_management_info(sheet: openpyxl.worksheet.worksheet.Worksheet):
    unmerge_whole_sheet(sheet)
    sheet.delete_rows(1, 2)
    sheet.cell(row=1, column=1).value = "PARAMETERS"
    create_country_header(sheet, 2, has_third_column=True)
    create_country_header(sheet, 5, has_third_column=True)
    create_country_header(sheet, 8, has_third_column=True)
    remove_empty_rows(sheet)
    sheet.delete_rows(2, 1)
    # make cell bold
    sheet.cell(row=43, column=1).font = openpyxl.styles.fonts.Font(bold=True)
    first_column_only_rows = find_all_sheet_rows_with_value_in_only_first_column(sheet)
    first_column_bolded_rows = find_all_sheet_rows_with_bolded_value_in_first_column(sheet)
    copy_first_only_values_to_following_rows(sheet, first_column_only_rows, first_column_bolded_rows)
    delete_rows(first_column_only_rows, sheet)


def prep_specific_sheet(workbook: openpyxl.Workbook, sheet_name: str):
    switcher = {
        'Balance Sheet': prep_balance_sheet,
        'Income Statement': prep_income_statement,
        'Management Info': prep_management_info,
    }
    prep_func = switcher.get(sheet_name, None)
    if prep_func is not None:
        prep_func(workbook[sheet_name])
    else:
        print(f"Sheet {sheet_name} will not be processed")


def prep_files(path):
    files = os.listdir(path)
    files.remove('processed')
    for file in files:
        if file != "START.xlsx":
            workbook = openpyxl.open(path + "\\" + file)
            workbook.remove(workbook.worksheets[0])
            for sheet in workbook.worksheets:
                prep_specific_sheet(workbook, sheet.title)
            workbook.save(path + "\\" + 'processed' + "\\" + file)
            workbook.close()
            print("Done with " + file)
        else:
            os.rename(path + "\\" + file, path + "\\" + 'processed' + "\\" + file)


def process_standard_cost(key, df):
    areas = ["U.S. PC 2", "EC/EU PC 2", "BRAZIL PC 2"]

    plants_number = len(df[df["PARAMETERS"].str.contains('PL\(')])

    standard_rows = []
    unit_rows = []
    for i in range(0, plants_number):
        pl_i_standard_row = df.loc[df["PARAMETERS"] == f"MANUFACTURING COST ANALYSIS PL({i+1}) STANDARD COST"]
        pl_i_row_index = pl_i_standard_row.index.item()
        pl_i_units_row = df.loc[pl_i_row_index + 1]
        standard_rows.append(pl_i_standard_row)
        unit_rows.append(pl_i_units_row)

    for area in areas:
        costs = sum(list(map(lambda row: int(row[area]), standard_rows)))
        amounts = sum(list(map(lambda row: int(row[area]), unit_rows)))
        variable_cost_per_unit = 0
        if amounts != 0:
            variable_cost_per_unit = costs / amounts
        if variable_cost_per_unit != 0:
            print(f"{key} {area} costs: {costs} amounts: {amounts} variable cost pre unit: {variable_cost_per_unit}")


def process_management_info(files_df):
    for key in files_df.keys():
        if key != "START":
            print(f"Management Info For {key} =======================================================================")
            df = files_df[key]["Management Info"]
            process_standard_cost(key, df)


if __name__ == '__main__':
    if validate_argv(sys.argv):
        path = sys.argv[1]
        should_process = sys.argv[2]
        if should_process == "True":
            prep_files(path)

        files_dict = read_excels_to_df(path)

        # Intra-co.purchase - סכום העברות כספים בהעברת מוצרים
        # Income statement -  INTRA-COMPANY - כמה נכנס לאיזור בעקבות העברת מוצרים
        process_management_info(files_dict)
        print("Done with all files")
